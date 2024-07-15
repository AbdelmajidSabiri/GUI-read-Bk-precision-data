import pyvisa
import time
import os
import pandas as pd
from openpyxl import load_workbook, Workbook

class Bkp8600(object):
    def __init__(self, resource=None):
        self.rm = pyvisa.ResourceManager()
        self.instr = None 
        self.instrument_found = False
        if resource:
            try:
                self.instr = self.rm.open_resource(resource, timeout=10000)
                self.instrument_found = True
            except pyvisa.VisaIOError:
                print(f"Error: Unable to open resource {resource}.")
        else:
            instruments = self.rm.list_resources()
            for r in instruments:
                try:
                    instr = self.rm.open_resource(r)
                    if instr.query("*IDN?").startswith("B&K Precision"):
                        self.instr = instr
                        self.instrument_found = True
                        break
                except pyvisa.VisaIOError:
                    pass
            if not self.instrument_found:
                print("No BK Precision instrument found.")

    def get_description(self):
        if self.instrument_found:
            return self.instr.query("*IDN?")

    def get_current(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:CURRent?"))

    def get_voltage(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:VOLTage?"))
    
    def get_resistance(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:RESistance?"))

    def get_power(self):
        if self.instrument_found:
            return float(self.instr.query(":MEASure:POWer?"))

    def initialize(self):
        if self.instrument_found:
            self.instr.write("SYSTem:REMote")
            self.instr.write("INPut OFF")
            self.instr.write("*RST")
            self.instr.write("*CLS")
            self.instr.write("*SRE 0")
            self.instr.write("*ESE 0")

    def set_current(self, current):
        if self.instrument_found:
            self.instr.write("INPut ON")
            self.instr.write("FUNC CURRent")
            self.instr.write(f"CURRent {current}")

    
    def set_voltage(self, voltage):
        if self.instrument_found:
            self.instr.write("INPut ON")  # Turn on the load
            self.instr.write("FUNC VOLTage")  # Set the function mode to Constant Voltage (CV)
            self.instr.write(f"VOLTage {voltage}") 

    def reset_to_manual(self):
        if self.instrument_found:
            self.instr.write("INPut OFF")
            self.instr.write("SYSTem:LOCal")

        
if __name__ == '__main__':
    try:
        last = Bkp8600()

        last.initialize()

        # Perform operations (e.g., setting current and measuring values)
        for n in range(0, 200):
            current_value = n / 100.0
            print(f"Setting current to: {current_value} A")
            last.set_current(current_value)
            time.sleep(1)
            print(f"Current: {last.get_current()} A")
            print(f"Voltage: {last.get_voltage()} V")



    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Reset to local mode and close the connection
        last.reset_to_manual()
        del last
def Add_current(AddedValue, filename="output.xlsx"):
    # Initialize Bkp8600 object and measure current and voltage
    last = Bkp8600()
    last.initialize()
    
    measured_current = last.get_current()
    voltage = last.get_voltage()
    last.set_current(AddedValue)
    
    # Create a DataFrame with the new data
    df = pd.DataFrame([[measured_current, voltage, AddedValue,None]], 
                      columns=["Measured Current", "Voltage", "Added Current","Added voltage"])
    
    # Load the existing workbook or create a new one if it doesn't exist
    try:
        book = load_workbook(filename)
    except FileNotFoundError:
        book = Workbook()
    
    # Select the active worksheet (you can modify this as per your requirement)
    sheet_name = 'Sheet1'
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    
    sheet = book[sheet_name]
    
    # Append data to the worksheet
    max_row = sheet.max_row
    for row in df.values.tolist():
        sheet.append(row)
    
    # Save the workbook
    book.save(filename)
    book.close()  # Close the workbook
    
    print(f"Data appended to {filename} successfully.")
    
    # Reset Bkp8600 object
    last.reset_to_manual()
    del last

def Add_voltage(AddedValue, filename="output.xlsx"):
    # Initialize Bkp8600 object and measure current and voltage
    last = Bkp8600()
    last.initialize()
    
    measured_current = last.get_current()
    voltage = last.get_voltage()
    last.set_voltage(AddedValue)
    
    # Create a DataFrame with the new data
    df = pd.DataFrame([[measured_current, voltage, None, AddedValue]], 
                      columns=["Measured Current", "Voltage", "Added Current","Added Voltage"])
    
    # Load the existing workbook or create a new one if it doesn't exist
    try:
        book = load_workbook(filename)
    except FileNotFoundError:
        book = Workbook()
    
    # Select the active worksheet (you can modify this as per your requirement)
    sheet_name = 'Sheet1'
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    
    sheet = book[sheet_name]
    
    # Append data to the worksheet
    max_row = sheet.max_row
    for row in df.values.tolist():
        sheet.append(row)
    
    # Save the workbook
    book.save(filename)
    book.close()  # Close the workbook
    
    print(f"Data appended to {filename} successfully.")
    
    # Reset Bkp8600 object
    last.reset_to_manual()
    del last